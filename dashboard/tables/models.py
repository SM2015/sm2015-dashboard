# coding: utf-8
import re
from openpyxl import load_workbook
from django.db import models
from core.models import Country, Language
from datetime import date


class Quarter(models.Model):
    name = models.CharField(max_length=7, default='')

    @classmethod
    def filter_objects_by_year(self, obj_list, year):
        new_list = []
        for obj in obj_list:
            obj_year = int(obj.quarter.name[:4])
            if obj_year == int(year):
                new_list.append(obj)
        return new_list

    @classmethod
    def get_actual_quarter(cls):
        today = date.today()
        if today.month / 3 <= 1:
            number_quarter = 1
        elif today.month / 3 <= 2:
            number_quarter = 2
        elif today.month / 3 <= 3:
            number_quarter = 3
        elif today.month / 3 <= 4:
            number_quarter = 4
        return cls.get_quarter_by_number(today.year, number_quarter)

    @classmethod
    def get_quarter_by_number(cls, year, number):
        quarter_name = "{0}Q{1}".format(year, number)
        try:
            quarter = cls.objects.get(name=quarter_name)
            return quarter
        except Quarter.DoesNotExist:
            return None

    @classmethod
    def normalize_name(cls, name):
        if name:
            if type(name) is int:
                name = "{0}".format(name)
            else:
                # Caso: 2012QI
                name = name.lower() \
                           .replace("iii", "3") \
                           .replace("ii", "2") \
                           .replace("i", "1")

                # Casos: "Q1 2012" ou "T1 2012"
                try:
                    if (name[0].lower() == 'q' or name[0].lower() == 't') and int(name[1]):
                        name = name.replace(' ', '')
                        name = "{0}{1}{2}".format(name[2:], name[0], name[1])
                except ValueError:
                    pass

            return name.upper()
        else:
            return None

    def __unicode__(self):
        return self.name

    class Meta:
        ordering = ['name']


class AvanceFisicoFinanciero(models.Model):
    country = models.ForeignKey(Country)
    language = models.ForeignKey(Language, default=1)

    fecha_de_actualizacion = models.DateField(null=True, default=None, blank=True)
    avance_fisico_planificado = models.FloatField(null=True, blank=True, default=None)
    avance_financiero_planificado = models.FloatField(null=True, blank=True, default=None)
    avance_fisico_real = models.FloatField(null=True, blank=True, default=None)
    avance_financiero_actual = models.FloatField(null=True, blank=True, default=None)
    avances_fisicos_original_programado = models.FloatField(null=True, blank=True, default=None)
    avances_financieros_original_programado = models.FloatField(null=True, blank=True, default=None)
    monto_comprometido = models.FloatField(null=True, blank=True, default=None)
    monto_desembolsado = models.FloatField(null=True, blank=True, default=None)

    alerta = models.TextField(null=True, blank=True, default=None)
    recomendacion = models.TextField(null=True, blank=True, default=None)
    upcoming_policy_dialogue_events = models.TextField(null=True, blank=True, default='')

    @classmethod
    def upload_excel(cls, uploaded_file, sheet_lang):
        wb = load_workbook(uploaded_file, data_only=True)
        sheets = [
            {'country': Country.objects.get(slug='belize'), 'sheet': wb.get_sheet_by_name('Belize')},
            {'country': Country.objects.get(slug='costa-rica'), 'sheet': wb.get_sheet_by_name('Costa Rica')},
            {'country': Country.objects.get(slug='el-salvador'), 'sheet': wb.get_sheet_by_name('El Salvador')},
            {'country': Country.objects.get(slug='guatemala'), 'sheet': wb.get_sheet_by_name('Guatemala')},
            {'country': Country.objects.get(slug='honduras'), 'sheet': wb.get_sheet_by_name('Honduras')},
            {'country': Country.objects.get(slug='mexico'), 'sheet': wb.get_sheet_by_name('Mexico')},
            {'country': Country.objects.get(slug='nicaragua'), 'sheet': wb.get_sheet_by_name('Nicaragua')},
            {'country': Country.objects.get(slug='panama'), 'sheet': wb.get_sheet_by_name('Panama')}
        ]

        language = Language.objects.get(acronym=sheet_lang)

        for sheet in sheets:
            real_sheet = sheet.get('sheet')
            country = sheet.get('country')

            try:
                fecha_de_actualizacion = real_sheet.rows[0][2].value.date()
            except AttributeError:
                fecha_de_actualizacion = None

            try:
                alerta = real_sheet.rows[0][11].value
            except IndexError:
                alerta = ''

            try:
                recomendacion = real_sheet.rows[1][11].value
            except IndexError:
                recomendacion = ''

            cls.objects.create(language = language,
                               country = country,
                               fecha_de_actualizacion = fecha_de_actualizacion,
                               avance_fisico_planificado = real_sheet.rows[0][4].value,
                               avance_financiero_planificado = real_sheet.rows[0][6].value,
                               avance_fisico_real = real_sheet.rows[1][4].value,
                               avance_financiero_actual = real_sheet.rows[1][6].value,
                               avances_fisicos_original_programado = real_sheet.rows[0][9].value,
                               avances_financieros_original_programado = real_sheet.rows[1][9].value,
                               monto_comprometido = real_sheet.rows[1][7].value,
                               monto_desembolsado = real_sheet.rows[1][2].value,
                               alerta = alerta,
                               recomendacion = recomendacion)

    @classmethod
    def get_editable_fields(cls):
        return ('fecha_de_actualizacion', 'avance_fisico_planificado', 'avance_financiero_planificado',
        'avance_fisico_real', 'avance_financiero_actual', 'avances_fisicos_original_programado',
        'avances_financieros_original_programado', 'monto_desembolsado', 'monto_comprometido',
        'alerta', 'recomendacion', 'upcoming_policy_dialogue_events')

    def __unicode__(self):
        return self.country.name

class Audiencia(models.Model):
    language = models.ForeignKey(Language, default=1)
    name = models.CharField(max_length=100)

    def __unicode__(self):
        return self.name

class EstadoActual(models.Model):
    language = models.ForeignKey(Language, default=1)
    name = models.CharField(max_length=100)

    def __unicode__(self):
        return self.name

class Hito(models.Model):
    country = models.ForeignKey(Country)
    language = models.ForeignKey(Language, default=1)
    quarter = models.ForeignKey(Quarter)

    indicador_de_pago = models.CharField(max_length=500, null=True, blank=True, default=None)
    hito = models.CharField(max_length=500, null=True, blank=True, default=None)
    audiencia = models.ManyToManyField(Audiencia)
    estado_actual = models.ForeignKey(EstadoActual, null=True, blank=True, default=None)
    alerta_notas = models.TextField(null=True, blank=True)
    recomendacion = models.TextField(null=True, blank=True)
    acuerdo = models.TextField(null=True, blank=True)
    actividad_en_poa = models.CharField(max_length=500, null=True, blank=True, default=None)

    @classmethod
    def upload_excel(cls, uploaded_file, sheet_lang):
        wb = load_workbook(uploaded_file, data_only=True)
        sheets = [
            {'country': Country.objects.get(slug='belize'), 'sheet': wb.get_sheet_by_name('Belize') },
            {'country': Country.objects.get(slug='costa-rica'), 'sheet': wb.get_sheet_by_name('Costa Rica') },
            {'country': Country.objects.get(slug='el-salvador'), 'sheet': wb.get_sheet_by_name('El Salvador') },
            {'country': Country.objects.get(slug='guatemala'), 'sheet': wb.get_sheet_by_name('Guatemala') },
            {'country': Country.objects.get(slug='honduras'), 'sheet': wb.get_sheet_by_name('Honduras') },
            {'country': Country.objects.get(slug='mexico'), 'sheet': wb.get_sheet_by_name('Mexico') },
            {'country': Country.objects.get(slug='nicaragua'), 'sheet': wb.get_sheet_by_name('Nicaragua') },
            {'country': Country.objects.get(slug='panama'), 'sheet': wb.get_sheet_by_name('Panama') }
        ]

        language = Language.objects.get(acronym=sheet_lang)

        for sheet in sheets:
            real_sheet = sheet.get('sheet')
            country = sheet.get('country')

            for row in real_sheet.rows:
                if row[0].row >= 6 and (row[1].value or row[2].value):
                    if country.slug in ['el-salvador', 'mexico', 'costa-rica']:
                        indicador_de_pago_str = row[1].value
                        hito_str = row[3].value
                        quarter_str = row[4].value
                        audiencias_str = row[5].value
                        estado_actual_str = row[6].value
                        alerta_notas_str = row[7].value
                        recomendacion_str = row[8].value
                        acuerdo_str = row[9].value
                        actividad_en_poa_str = row[10].value
                    elif country.slug in ['belize']:
                        indicador_de_pago_str = row[1].value
                        hito_str = row[2].value
                        quarter_str = row[3].value
                        audiencias_str = row[5].value
                        estado_actual_str = row[4].value
                        alerta_notas_str = row[6].value
                        recomendacion_str = row[7].value
                        acuerdo_str = row[8].value
                        actividad_en_poa_str = row[9].value
                    else:
                        indicador_de_pago_str = row[1].value
                        hito_str = row[2].value
                        quarter_str = row[3].value
                        audiencias_str = row[4].value
                        estado_actual_str = row[5].value
                        alerta_notas_str = row[6].value
                        recomendacion_str = row[7].value
                        acuerdo_str = row[8].value
                        actividad_en_poa_str = row[9].value

                    try:
                        if estado_actual_str.lower() == 'completed' or estado_actual_str.lower() == 'cumplido':
                            estado_actual = EstadoActual.objects.get(name='Cumplido')
                        elif estado_actual_str.lower() == 'in progress' or estado_actual_str.lower() == 'en proceso':
                            estado_actual = EstadoActual.objects.get(name='En proceso')
                        elif estado_actual_str.lower() == 'delayed' or estado_actual_str.lower() == 'retrasado':
                            estado_actual = EstadoActual.objects.get(name='Retrasado')
                        else:
                            estado_actual = None
                    except AttributeError:
                        estado_actual = None

                    try:
                        quarter = Quarter.objects.get(name=Quarter.normalize_name(quarter_str))
                    except:
                        quarter = Quarter.objects.create(name=Quarter.normalize_name(quarter_str))

                    hito = cls.objects.create(
                        language = language,
                        country = country,
                        indicador_de_pago = indicador_de_pago_str,
                        hito = hito_str,
                        quarter = quarter,
                        estado_actual = estado_actual,
                        alerta_notas = alerta_notas_str,
                        recomendacion = recomendacion_str,
                        acuerdo = acuerdo_str,
                        actividad_en_poa = actividad_en_poa_str
                    )

                    if audiencias_str:
                        text_audiencias = audiencias_str.replace(' ', '').split(',')
                        for i in xrange(0, len(text_audiencias)):
                            audiencia_str = text_audiencias[i].replace('Country', 'Pais').replace('Donors', 'Donantes')
                            text_audiencias[i] = audiencia_str

                        audiencias = Audiencia.objects.filter(name__in=text_audiencias)
                        for audiencia in audiencias:
                            hito.audiencia.add(audiencia)
                        hito.save()

    @classmethod
    def get_editable_fields(cls):
        return ('estado_actual', 'alerta_notas', 'recomendacion', 'acuerdo', 'actividad_en_poa')

    def __unicode__(self):
        return self.country.name

class UcMilestone(models.Model):
    quarter = models.ForeignKey(Quarter)
    language = models.ForeignKey(Language, default=1)

    objective = models.CharField(max_length=300, null=True, blank=True, default=None)
    coordination_unit_milestone = models.CharField(max_length=500, null=True, blank=True, default=None)
    status = models.CharField(max_length=200, null=True, blank=True, default=None)
    observation = models.CharField(max_length=500, null=True, blank=True, default=None)

    @classmethod
    def get_dates(cls):
        dates = []
        for row in cls.objects.values('quarter__name').distinct():
            if row['quarter__name']:
                year = row['quarter__name'][:4]
                if int(year) not in dates:
                    dates.append(int(year))
        return dates

    @classmethod
    def upload_excel(cls, uploaded_file, sheet_name, sheet_lang):
        wb = load_workbook(uploaded_file, data_only=True)
        sheet = wb.get_sheet_by_name(sheet_name)
        language = Language.objects.get(acronym=sheet_lang)

        date_field = None
        test_find_year = re.search('(2010|2011|2012|2013|2014|2015)',
                                   sheet_name)
        if test_find_year:
            year = int(test_find_year.group(0))
            date_field = date(year, 1, 1)

        for row in sheet.rows:
            if not row[0].row == 1 and row[0].value:
                try:
                    quarter = Quarter.objects.get(name=Quarter.normalize_name(row[2].value))
                except:
                    quarter = Quarter.objects.create(name=Quarter.normalize_name(row[2].value))

                cls.objects.create(
                    date=date_field,
                    language=language,
                    objective=row[0].value,
                    coordination_unit_milestone=row[1].value,
                    quarter=quarter,
                    status=row[3].value,
                    observation=row[4].value
                )

    @classmethod
    def get_editable_fields(cls):
        return ('objective', 'coordination_unit_milestone', 'quarter', 'status', 'observation')

    def __unicode__(self):
        return self.coordination_unit_milestone

class Objective(models.Model):
    language = models.ForeignKey(Language, default=1)
    objective = models.CharField(max_length=200, null=True, blank=True, default=None)
    
    def __unicode__(self):
        return self.objective

class Sm2015Milestone(models.Model):
    objective = models.ForeignKey(Objective, null=True, blank=True, default=None)
    language = models.ForeignKey(Language, default=1)

    date = models.DateField(null=True, default=None)
    hitos = models.CharField(max_length=500, null=True, blank=True, default=None)
    status = models.CharField(max_length=200, null=True, blank=True, default=None)
    observation = models.CharField(max_length=500, null=True, blank=True, default=None)

    @classmethod
    def get_dates(cls):
        dates = []
        for row in cls.objects.values('date').distinct():
            if row['date']:
                dates.append(row['date'].year)
        return dates

    @classmethod
    def upload_excel(cls, uploaded_file, sheet_name, sheet_lang):
        wb = load_workbook(uploaded_file, data_only=True)
        sheet = wb.get_sheet_by_name(sheet_name)
        language = Language.objects.get(acronym=sheet_lang)

        date_field = None
        test_find_year = re.search('(2010|2011|2012|2013|2014|2015)',
                                   sheet_name)
        if test_find_year:
            year = int(test_find_year.group(0))
            date_field = date(year, 1, 1)

        for row in sheet.rows:
            if not row[0].row == 1 and row[1].value:
                if row[0].value:
                    try:
                        objective = Objective.objects.get(objective=
                                                          row[0].value)
                    except:
                        objective = Objective.objects.create(objective=
                                                             row[0].value,
                                                             language=language)

                cls.objects.create(
                    date=date_field,
                    language=language,
                    objective=objective,
                    hitos=row[1].value,
                    status=row[2].value,
                    observation=row[3].value
                )

    @classmethod
    def get_editable_fields(cls):
        return ('hitos', 'status', 'observation')

    def __unicode__(self):
        return self.hitos

class GrantsFinancesType(models.Model):
    name = models.CharField(max_length=200, default='')
    uuid = models.CharField(max_length=200, default='')

    def __unicode__(self):
        return self.name

class GrantsFinancesOrigin(models.Model):
    name = models.CharField(max_length=200, default='')
    uuid = models.CharField(max_length=200, default='')

    def __unicode__(self):
        return self.name

class GrantsFinancesFields(models.Model):
    name = models.CharField(max_length=200, default='')
    field_type = models.ForeignKey(GrantsFinancesType, null=True)
    field_origin = models.ForeignKey(GrantsFinancesOrigin, null=True)

    @classmethod
    def get_editable_fields(cls):
        return ('name', 'field_origin', 'field_type')

    def __unicode__(self):
        return self.name

    class Meta:
        verbose_name_plural = u'Grants & Finances Fields'
        verbose_name = u'Grants & Finances Fields'

class GrantsFinances(models.Model):
    quarter = models.ForeignKey(Quarter)
    field = models.ForeignKey(GrantsFinancesFields, null=True)
    value = models.FloatField(default=0)

    @classmethod
    def get_editable_fields(cls):
        return ('value',)

    @classmethod
    def get_periods(cls):
        periods = []
        for quarter_dict in GrantsFinances.objects.values('quarter__name').order_by('quarter__name').distinct():
            periods.append(quarter_dict['quarter__name'])
        return periods

    @classmethod
    def get_accumulated(cls, uuid_origin, uuid_type):
        accumulated = 0
        field = GrantsFinancesFields.objects.filter(field_type__uuid=uuid_type).filter(field_origin__uuid=uuid_origin)
        for grant in cls.objects.filter(field=field):
            accumulated += grant.value
        return accumulated

    @classmethod
    def upload_excel(cls, uploaded_file):
        wb = load_workbook(uploaded_file, data_only=True)
        sheet = wb.get_sheet_by_name('D.1.1.')

        bmgf_origin = GrantsFinancesOrigin.objects.get(uuid="GRANTS_ORIGIN_BMFG")
        icss_origin = GrantsFinancesOrigin.objects.get(uuid="GRANTS_ORIGIN_ICSS")
        gos_origin = GrantsFinancesOrigin.objects.get(uuid="GRANTS_ORIGIN_GOS")
        korean_origin = GrantsFinancesOrigin.objects.get(uuid="GRANTS_ORIGIN_KOREAN")

        real_type =  GrantsFinancesType.objects.get(uuid="GRANTS_TYPE_REAL")
        expected_type =  GrantsFinancesType.objects.get(uuid="GRANTS_TYPE_EXPECTED")

        columns_index = {
                'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7,
                'I': 8, 'J': 9, 'K': 10, 'L': 11, 'M': 12, 'N': 13, 'O': 14, 
                'P': 15, 'Q': 16, 'R': 17, 'S': 18, 'T': 19, 'U': 20, 'V': 21, 
                'W': 22, 'X': 23, 'Y': 24, 'Z': 25
        }
        map_rows = {
            5: {'field': GrantsFinancesFields.objects.get(field_origin=bmgf_origin, field_type=expected_type)},
            9: {'field': GrantsFinancesFields.objects.get(field_origin=icss_origin, field_type=expected_type)},       
            11: {'field': GrantsFinancesFields.objects.get(field_origin=gos_origin, field_type=expected_type)},       
            13: {'field': GrantsFinancesFields.objects.get(field_origin=korean_origin, field_type=expected_type)},       
            18: {'field': GrantsFinancesFields.objects.get(field_origin=bmgf_origin, field_type=real_type)},       
            22: {'field': GrantsFinancesFields.objects.get(field_origin=icss_origin, field_type=real_type)},       
            24: {'field': GrantsFinancesFields.objects.get(field_origin=gos_origin, field_type=real_type)},      
            26: {'field': GrantsFinancesFields.objects.get(field_origin=korean_origin, field_type=real_type)}       
        }
        
        period_row = sheet.rows[2]
        for row in sheet.rows:
            try:
                map_row = map_rows[row[0].row]
                for cell in row:
                    if cell.column not in ['A', 'B'] and (cell.value or cell.value == 0):
                        try:
                            quarter = Quarter.objects.get(name=Quarter.normalize_name(period_row[columns_index[cell.column]].value))
                        except:
                            quarter = Quarter.objects.create(name=Quarter.normalize_name(period_row[columns_index[cell.column]].value))

                        cls.objects.create(
                            quarter=quarter,
                            field=map_row['field'],
                            value=cell.value * 1000000
                        )
            except KeyError:
                continue

    def __unicode__(self):
        return self.field.name

    class Meta:
        verbose_name_plural = u'Grants & Finances'
        verbose_name = u'Grants & Finances'


class Operation(models.Model):
    country = models.ForeignKey(Country)

    number = models.CharField(max_length=100, default='')
    name = models.CharField(max_length=100, default='')
    executing_agency = models.CharField(max_length=100, default='')
    benefitted_population = models.CharField(max_length=400, default='')
    starting_date = models.DateField()
    finish_date = models.DateField()
    objectives_progress = models.TextField()
    key_results_expected = models.TextField()

    def __unicode__(self):
        return self.name


class OperationTotalInvestment(models.Model):
    operation = models.ForeignKey(Operation)

    investment_tranche_first_operation = models.FloatField()
    investment_tranche_second_operation = models.FloatField()
    investment_tranche_third_operation = models.FloatField()
    counterpart_tranche_first_operation = models.FloatField()
    counterpart_tranche_second_operation = models.FloatField()
    counterpart_tranche_third_operation = models.FloatField()
    cost_of_the_operation_first_operation = models.FloatField()
    cost_of_the_operation_second_operation = models.FloatField()
    cost_of_the_operation_third_operation = models.FloatField()
    performance_tranche_first_operation = models.FloatField()
    performance_tranche_second_operation = models.FloatField()
    performance_tranche_third_operation = models.FloatField()

    def __unicode__(self):
        return self.operation.name


class OperationZones(models.Model):
    operation = models.ForeignKey(Operation)

    name = models.CharField(max_length=100)
    slug = models.SlugField(max_length=100)
    latlng = models.CharField(max_length=100, default='')

    def __unicode__(self):
        return self.name


class LifeSaveField(models.Model):
    name = models.CharField(max_length=500, default='', null=True)
    abbr = models.CharField(max_length=200, default='', null=True)

    @classmethod
    def get_editable_fields(cls):
        return ('name', 'abbr')

    def __unicode__(self):
        return self.name

    class Meta:
        verbose_name_plural = u'LiSTs Fields'
        verbose_name = u'LiST Fields'

class LifeSave(models.Model):
    country = models.ForeignKey(Country)
    field = models.ForeignKey(LifeSaveField, null=True)

    percentage = models.FloatField(default=0)
    number_saving = models.IntegerField(default=0)

    @classmethod
    def upload_excel(cls, uploaded_file):
        wb = load_workbook(uploaded_file, data_only=True)
        sheets = [
            {'country': Country.objects.get(slug='costa-rica'), 'sheet': wb.get_sheet_by_name('LiST2 CR') },
            {'country': Country.objects.get(slug='nicaragua'), 'sheet': wb.get_sheet_by_name('LiST2 NI') }
        ]
        
        for sheet in sheets:
            real_sheet = sheet.get('sheet')
            country = sheet.get('country')
            
            for row in real_sheet.rows:
                if row[0].row >= 3 and row[0].row <= 21:
                    try:
                        field = LifeSaveField.objects.get(name=row[0].value, abbr=row[1].value)
                    except:
                        field = LifeSaveField.objects.create(name=row[0].value, abbr=row[1].value)

                    number_saving = re.sub("\D", "", str(row[3].value))
                    percentage = re.sub("\D", "", str(row[2].value))
                    if percentage == '':
                        percentage = 0
                    if number_saving == '':
                        number_saving = 0

                    cls.objects.create(
                        country = country,
                        field = field,
                        percentage = int(percentage),
                        number_saving = int(number_saving)
                    )

    def __unicode__(self):
        return self.country.name

    @classmethod
    def get_editable_fields(cls):
        return ('percentage', 'number_saving')

    class Meta:
        verbose_name_plural = u'LiSTs'
        verbose_name = u'LiST'

class CountryDisbursementCharger(models.Model):
    name = models.CharField(max_length=200)

    def __unicode__(self):
        return self.name

class CountryDisbursement(models.Model):
    country = models.ForeignKey(Country, default=None, null=True)
    operation = models.ForeignKey(Operation, default=None, null=True)
    quarter = models.ForeignKey(Quarter)
    charger = models.ForeignKey(CountryDisbursementCharger, default=None, null=True)

    year = models.CharField(max_length=4, default='', null=True)
    date = models.DateField(default=None, null=True)
    description = models.CharField(max_length=300, default='', null=True)
    amount = models.FloatField(default=0)

    def __unicode__(self):
        return self.country.name

    @classmethod
    def get_editable_fields(cls):
        return ('year', 'quarter', 'date', 'description', 'amount')

    @classmethod
    def upload_excel(cls, uploaded_file):
        wb = load_workbook(uploaded_file, data_only=True)
        sheets = [
            {'country': Country.objects.get(slug='belize'), 'sheet': wb.get_sheet_by_name('Belize'), 'row_starts': 64, 'row_ends': 99 },
            {'country': Country.objects.get(slug='costa-rica'), 'sheet': wb.get_sheet_by_name('Costa Rica'), 'row_starts': 125, 'row_ends': 160 },
            {'country': Country.objects.get(slug='el-salvador'), 'sheet': wb.get_sheet_by_name('El Salvador'), 'row_starts': 69, 'row_ends': 104 },
            {'country': Country.objects.get(slug='guatemala'), 'sheet': wb.get_sheet_by_name('Guatemala'), 'row_starts': 55, 'row_ends': 90 },
            {'country': Country.objects.get(slug='honduras'), 'sheet': wb.get_sheet_by_name('Honduras'), 'row_starts': 65, 'row_ends': 100 },
            {'country': Country.objects.get(slug='mexico'), 'sheet': wb.get_sheet_by_name('Mexico'), 'row_starts': 121 , 'row_ends': 156 },
            {'country': Country.objects.get(slug='nicaragua'), 'sheet': wb.get_sheet_by_name('Nicaragua'), 'row_starts': 68, 'row_ends': 103 },
            {'country': Country.objects.get(slug='panama'), 'sheet': wb.get_sheet_by_name('Panama'), 'row_starts': 53, 'row_ends': 88 }
        ]
        
        for sheet in sheets:
            real_sheet = sheet.get('sheet')
            country = sheet.get('country')
            
            for row in real_sheet.rows:
                if row[0].row >= sheet.get('row_starts') and row[0].row <= sheet.get('row_ends'):
                    try:
                        operation = Operation.objects.get(name=row[2].value)
                    except Operation.DoesNotExist:
                        operation = None

                    try:
                        charger = CountryDisbursementCharger.objects.get(name=row[1].value)
                    except CountryDisbursementCharger.DoesNotExist:
                        charger = CountryDisbursementCharger.objects.create(name=row[1].value)

                    try:
                        quarter = Quarter.objects.get(name=Quarter.normalize_name(row[3].value))
                    except:
                        quarter = Quarter.objects.create(name=Quarter.normalize_name(row[3].value))

                    try:
                        date = row[4].value.date()
                    except AttributeError:
                        date = None
            
                    amount = re.sub("\D", "", str(row[6].value))
                    if amount == '':
                        amount = 0

                    cls.objects.create(
                        country = country,
                        operation = operation,
                        charger = charger,
                        year = row[0].value,
                        quarter = quarter,
                        date = date,
                        description = row[5].value,
                        amount = amount
                    )


class CountryOperationIT(models.Model):
    country = models.ForeignKey(Country, default=None, null=True)
    it = models.FloatField(default=0)

    def __unicode__(self):
        return self.country.name

    @classmethod
    def get_editable_fields(cls):
        return ('it',)


class CountryOperation(models.Model):
    country = models.ForeignKey(Country, default=None, null=True)
    quarter = models.ForeignKey(Quarter)

    it_disbursements_planned = models.FloatField(default=0)
    it_disbursements_actual = models.FloatField(default=0)
    it_execution_planned = models.FloatField(default=0)
    it_execution_actual = models.FloatField(default=0)

    def __unicode__(self):
        return self.country.name

    @classmethod
    def get_editable_fields(cls):
        return ('it_disbursements_planned', 'it_disbursements_actual',
                'it_execution_planned', 'it_execution_actual')

    @classmethod
    def get_last_quarter(cls, country):
        return cls.objects.filter(country=country).order_by('-quarter')[0]

    @classmethod
    def get_actual_quarter(cls, country):
        today = date.today()
        if today.month / 3 <= 1:
            number_quarter = 1
        elif today.month / 3 <= 2:
            number_quarter = 2
        elif today.month / 3 <= 3:
            number_quarter = 3
        elif today.month / 3 <= 4:
            number_quarter = 4

        quarter = Quarter.get_quarter_by_number(today.year, number_quarter)
        return cls.objects.filter(country=country, quarter=quarter)[0]

    @classmethod
    def get_table_to_show(cls, country=None, fields=None, with_total=False, until_quarter=None):
        table = []
        if country:
            countries = [country]
        else:
            countries = CountryOperation.objects.values('country__name').distinct()
        quarters = CountryOperation.objects.values('quarter__name').order_by('quarter__name') \
                                           .distinct()

        if until_quarter:
            for i_quarter in xrange(0, len(quarters)):
                quarter = quarters[i_quarter]
                if quarter['quarter__name'] == until_quarter.name:
                    quarters = quarters[:i_quarter+1]
                    break

        fields = [{'field': 'it_disbursements_planned',
                   'name': 'Planned (PD)',
                   'parent_name': 'IT Disbursements'},
                  {'field': 'it_disbursements_actual',
                   'name': 'Actual (AD)',
                   'parent_name': 'IT Disbursements'},
                  {'field': 'it_execution_planned',
                   'name': 'Planned (FP)',
                   'parent_name': 'IT Execution'},
                  {'field': 'it_execution_actual',
                   'name': 'Actual (AE)',
                   'parent_name': 'IT Execution'}]

        if with_total:
            total = [{'field': fields[0],
                      'country': "All Countries",
                      'it': 0,
                      'values': [0 for i in xrange(0, len(quarters))]},
                     {'field': fields[1],
                      'country': "All Countries",
                      'it': 0,
                      'values': [0 for i in xrange(0, len(quarters))]},
                     {'field': fields[2],
                      'country': "All Countries",
                      'it': 0,
                      'values': [0 for i in xrange(0, len(quarters))]},
                     {'field': fields[3],
                      'country': "All Countries",
                      'it': 0,
                      'values': [0 for i in xrange(0, len(quarters))]}]

        for country_name in countries:
            if not country:
                country_operation = Country.objects.get(name=country_name['country__name'])
            else:
                country_operation = country

            operation_it = CountryOperationIT.objects.get(country=country_operation)
            operation = CountryOperation.objects.filter(country=country_operation)
            country_rows = []
            for i in xrange(0, len(fields)):
                field = fields[i]
                row = {'field': field,
                       'country': country_operation.name,
                       'it_id': operation_it.id,
                       'it': operation_it.it / 1000000,
                       'values': []}
                if with_total:
                    total[i]['it'] += operation_it.it / 1000000

                for i_quarter in xrange(0, len(quarters)):
                    quarter = quarters[i_quarter]
                    operation_quarter = operation.get(quarter__name=quarter['quarter__name'])
                    value = getattr(operation_quarter, field['field'])
                    row['values'].append({'v': value, 'id': operation_quarter.id})

                    if with_total:
                        total[i]['values'][i_quarter] += value
                country_rows.append(row)
            table.extend(country_rows)
        if with_total:
            return table, quarters, total
        else:
            return table, quarters

    @classmethod
    def upload_excel(cls, uploaded_file):
        wb = load_workbook(uploaded_file, data_only=True)
        sheet = wb.get_sheet_by_name('Quarterly Performance')
        countries = {}

        for row in sheet.rows:
            country = None
            if row[0].row >= 5 and row[0].row <= 8:
                country = Country.objects.get(slug='mexico')
            elif row[0].row >= 9 and row[0].row <= 12:
                country = Country.objects.get(slug='belize')
            elif row[0].row >= 13 and row[0].row <= 16:
                country = Country.objects.get(slug='el-salvador')
            elif row[0].row >= 17 and row[0].row <= 20:
                country = Country.objects.get(slug='costa-rica')
            elif row[0].row >= 21 and row[0].row <= 24:
                country = Country.objects.get(slug='honduras')
            elif row[0].row >= 25 and row[0].row <= 28:
                country = Country.objects.get(slug='guatemala')
            elif row[0].row >= 29 and row[0].row <= 32:
                country = Country.objects.get(slug='nicaragua')
            elif row[0].row >= 33 and row[0].row <= 36:
                country = Country.objects.get(slug='panama')
            elif row[0].row == 4:
                quarters = {}
                for cell in row[5:16]:
                    quarters[cell.column] = cell.value

            if country:
                try:
                    CountryOperationIT.objects.get(country=country)
                except CountryOperationIT.DoesNotExist:
                    CountryOperationIT.objects.create(country=country, it=row[2].value * 1000000)

                if not countries.has_key(country.name):
                    countries[country.name] = {}

                if row[0].row in [5, 9, 13, 17, 21, 25, 29, 33]:
                    for cell in row[5:16]:
                        quarter = quarters[cell.column]
                        countries[country.name][quarter] = {}
                        countries[country.name][quarter]['country'] = country
                        countries[country.name][quarter]['it_disbursements_planned'] = cell.value * 1000000 or 0

                elif row[0].row in [6, 10, 14, 18, 22, 26, 30, 34]:
                    for cell in row[5:16]:
                        quarter = quarters[cell.column]
                        countries[country.name][quarter]['it_disbursements_actual'] = cell.value * 1000000 or 0

                elif row[0].row in [7, 11, 15, 19, 23, 27, 31, 35]:
                    for cell in row[5:16]:
                        quarter = quarters[cell.column]
                        countries[country.name][quarter]['it_execution_planned'] = cell.value * 1000000 or 0

                elif row[0].row in [8, 12, 16, 20, 24, 28, 32, 36]:
                    for cell in row[5:16]:
                        quarter = quarters[cell.column]
                        countries[country.name][quarter]['it_execution_actual'] = cell.value * 1000000 or 0

        for country_name in countries:
            quarters = countries[country_name]
            for quarter_name in quarters:
                row = quarters[quarter_name]

                try:
                    quarter = Quarter.objects.get(name=Quarter.normalize_name(quarter_name))
                except:
                    quarter = Quarter.objects.create(name=Quarter.normalize_name(quarter_name))

                cls.objects.create(country=row['country'],
                                   quarter=quarter,
                                   it_disbursements_planned=row['it_disbursements_planned'],
                                   it_disbursements_actual=row['it_disbursements_actual'],
                                   it_execution_planned=row['it_execution_planned'],
                                   it_execution_actual=row['it_execution_actual'])


class CountryDetailsIsech(models.Model):
    name = models.CharField(max_length=500, default='', null=True)
    numero = models.FloatField(default=None, null=True)

    def __unicode__(self):
        return self.name

class CountryDetailsLevel(models.Model):
    name = models.CharField(max_length=500, default='', null=True)

    def __unicode__(self):
        return self.name

class CountryDetailsPago(models.Model):
    name = models.CharField(max_length=500, default='', null=True)
    numero = models.IntegerField(default=None, null=True)

    def __unicode__(self):
        return self.name[:150]


class CountryDetails(models.Model):
    country = models.ForeignKey(Country, default=None, null=True)
    pago = models.ForeignKey(CountryDetailsPago, default=None, null=True)
    isech = models.ForeignKey(CountryDetailsIsech, default=None, null=True)
    level = models.ForeignKey(CountryDetailsLevel, default=None, null=True)

    location = models.CharField(max_length=300, default='', null=True)

    def __unicode__(self):
        return self.country.name

    @classmethod
    def get_editable_fields(cls):
        return ('pago', 'isech',
                'level', 'location')

    @classmethod
    def get_periods(cls, queryset=None):
        if not queryset:
            periods = []
            for quarter_dict in CountryDetailsValues.objects.values('quarter__name').order_by('quarter__name').distinct():
                periods.append(quarter_dict['quarter__name'])
        else:
            periods = queryset.values('countrydetailsvalues__quarter__name').distinct()
            periods = [v.get('countrydetailsvalues__quarter__name') for v in periods]

        return periods

    def get_instance_periods(self):
        values = self.values('countrydetailsvalues__quarter__name').distinct()
        values = [v.get('countrydetailsvalues__quarter__name') for v in values]
        return values

    @classmethod
    def upload_excel(cls, uploaded_file, sheet_name, sheet_country, **kw):
        wb = load_workbook(uploaded_file, data_only=True)
        sheet = wb.get_sheet_by_name(sheet_name)
        country = Country.objects.get(slug=sheet_country)
        for row in sheet.rows:
            if row[0].row == 6:
                quarters = []
                for row in row[6:]:
                    try:
                        quarter_name = row.value.strip()[::-1][:7][::-1]
                        quarter_normalized_name = Quarter.normalize_name(quarter_name)
                        # Verifica se a string se trata de um quarter
                        int(quarter_normalized_name[:4])
                        int(quarter_normalized_name[5])
                        quarter = Quarter.objects.get_or_create(name=quarter_normalized_name)[0]
                        if quarter not in quarters:
                            quarters.append(quarter)
                    except (ValueError, AttributeError):
                        continue

            elif row[0].row >= 7 and type(row[0].value) is int:
                try:
                    pago = CountryDetailsPago.objects.get(numero=row[0].value)
                except CountryDetailsPago.DoesNotExist:
                    pago = CountryDetailsPago.objects.create(numero=row[0].value,
                                                             name=row[1].value)

                try:
                    isech = CountryDetailsIsech.objects.get(numero=row[2].value)
                except CountryDetailsIsech.DoesNotExist:
                    isech = CountryDetailsIsech.objects.create(numero=row[2].value,
                                                               name=row[3].value)

                try:
                    level = CountryDetailsLevel.objects.get(name=row[4].value)
                except CountryDetailsLevel.DoesNotExist:
                    level = CountryDetailsLevel.objects.create(name=row[4].value)

                country_detail = cls.objects.create(country=country,
                                                    pago=pago,
                                                    isech=isech,
                                                    level=level,
                                                    location=row[5].value)

                last_value_column = 6
                for quarter in quarters:
                    cells_value = row[last_value_column:][:3]
                    numerador_reportado = cells_value[0].value
                    if type(numerador_reportado) not in [int, float]:
                        numerador_reportado = None

                    denominador_reportado = cells_value[1].value
                    if type(denominador_reportado) not in [int, float]:
                        denominador_reportado = None

                    CountryDetailsValues.objects.create(
                        country_detail=country_detail,
                        numerador_reportado=numerador_reportado,
                        denominador_reportado=denominador_reportado,
                        quarter=quarter
                    )
                    last_value_column += 3


class CountryDetailsValues(models.Model):
    country_detail = models.ForeignKey(CountryDetails)
    numerador_reportado = models.FloatField(default=None, null=True)
    denominador_reportado = models.FloatField(default=None, null=True)
    quarter = models.ForeignKey(Quarter)

    def __unicode__(self):
        return "{0} - {1}".format(self.country_detail.country,
                                  self.quarter)

    @classmethod
    def get_editable_fields(cls):
        return ('numerador_reportado', 'denominador_reportado',
                'quarter')
